# ----- Librerias ------

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, Toplevel, Label, Button
import pandas as pd
import cv2  # Para la captura de QR
from pyzbar.pyzbar import decode  # Decodificar QR
import os
import requests
import json
from dotenv import load_dotenv
import urllib3
import re
import numpy as np
import threading
import queue
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- Configuraciones -------------

# Deshabilitar las advertencias de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Cargar las variables del archivo .env
load_dotenv()

# Configuración de la API de GLPI
GLPI_URL = os.getenv("GLPI_URL")
USER_TOKEN = os.getenv("USER_TOKEN")
APP_TOKEN = os.getenv("APP_TOKEN")
PATH_EXCEL_NETWORK = os.getenv("PATH_EXCEL_NETWORK")
IP_CAM_URL = os.getenv("IP_CAM_URL")

# Ruta del archivo Excel
ruta_excel = PATH_EXCEL_NETWORK

class NetworkEquipment: 
    def __init__(self, root):
        self.crear_archivo_excel_con_hojas(ruta_excel, ["NetworkEquipment", "NetworkEquipment new"])
        self.lock_excel = threading.Lock()
        self.lock_glpi = threading.Lock()
        self.root = root
        self.root.title("GLPI Network Equipments Automator")
        self.root.geometry("600x600")
        self.style = ttk.Style()
        self.style.theme_use("clam")  # Puedes cambiar el tema a "clam", "alt", "default", "classic"
        self.configure_styles()
        self.create_widgets()
        try:
            session_token = self.obtener_token_sesion()
            if session_token:
                self.obtener_todos_los_network_equipment_glpi_a_excel(session_token)
        except Exception as e:
            print(f"Error al conectar con GLPI: {str(e)}")
            messagebox.showerror("Error", f"No se pudo conectar con GLPI, NO PODRAS REALIZAR SINCRONIZACIONES AUN: {str(e)}")

    def salir(self):
        root.destroy()  

    # --- Network Equipment --- 

    def configure_styles(self):
        # Estilo del marco
        self.style.configure("TFrame", background="#E0F7FA")

        # Estilo de etiquetas
        self.style.configure("TLabel", background="#E0F7FA", foreground="#01579B", font=("Montserrat", 12))
        self.style.configure("Header.TLabel", background="#01579B", foreground="#FFFFFF", font=("Montserrat", 16, "bold"))

        # Estilo de botones con bordes redondeados y efectos suaves
        self.style.configure("Rounded.TButton",
                            background="#0288D1",
                            foreground="#FFFFFF",
                            font=("Roboto", 12),
                            padding=(15, 10),  # Aumentar el padding para suavizar
                            borderwidth=2,
                            relief="flat")  # 'flat' para quitar bordes bruscos

        # Aplicar efecto hover (cuando el mouse está sobre el botón)
        self.style.map("Rounded.TButton",
                    background=[("active", "#0277BD")],  # Cambio de color al pasar el mouse
                    relief=[("pressed", "groove")])  # Suaviza el clic en el botón

    def create_widgets(self):
        # Menú
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Opciones", menu=file_menu)
        file_menu.add_command(label="Salir", command=self.root.quit)

        #help_menu = tk.Menu(menubar, tearoff=0)
        #menubar.add_cascade(label="Ayuda", menu=help_menu)
        #help_menu.add_command(label="Acerca de")

        # Pestañas
        tab_control = ttk.Notebook(self.root)
        tab_names = ["Registros Offline (Excel)", "Registros Online (GLPI)", "Excel -> GLPI (Sincronizacion Asincrona)"]
        frames = {}

        for name in tab_names:
            frames[name] = ttk.Frame(tab_control, padding="10")
            frames[name].pack(fill="both", expand=True)
            tab_control.add(frames[name], text=name)
            self.center_widgets(frames[name])  # Aplicar centrado global a cada pestaña

        tab_control.pack(expand=1, fill="both")

        # Excel
        ttk.Label(frames["Registros Offline (Excel)"], text="Registro de Network equipments a Excel", style="Header.TLabel").grid(row=0, column=0, pady=10)
        tk.Button(frames["Registros Offline (Excel)"], text="Registrar Network equipment a Excel", command=self.registro_offline).grid(row=1, column=0, padx=10, pady=5)
        tk.Button(frames["Registros Offline (Excel)"], text="Registrar multiples Network equipment a Excel", command=self.multiples_registros_offline).grid(row=2, column=0, padx=10, pady=5)

        # GLPI
        ttk.Label(frames["Registros Online (GLPI)"], text="Registro de Network equipments a GLPI", style="Header.TLabel").grid(row=0, column=0, pady=10)
        tk.Button(frames["Registros Online (GLPI)"], text="Registrar Network equipment a GLPI", command=self.registro_online).grid(row=1, column=0, padx=10, pady=5)
        tk.Button(frames["Registros Online (GLPI)"], text="Registrar multiples Network equipment a GLPI", command=self.multiples_registros_online).grid(row=2, column=0, padx=10, pady=5)

        # Excel a GLPI
        ttk.Label(frames["Excel -> GLPI (Sincronizacion Asincrona)"], text="Excel -> GLPI", style="Header.TLabel").grid(row=0, column=0, pady=10)
        tk.Button(frames["Excel -> GLPI (Sincronizacion Asincrona)"], text="Sincronizar Excel con GLPI", command=self.sincronizacion_asincrona).grid(row=1, column=0, padx=10, pady=5)

    def center_widgets(self, frame):
        # Configurar la columna 0 del frame para centrar elementos
        frame.columnconfigure(0, weight=1)

    ## --- Excel ---

    # Crear archivo Excel si no existe
    def crear_archivo_excel_con_hojas(self, ruta, hojas):
        if not os.path.exists(ruta):
            wb = Workbook()
            for hoja in hojas:
                ws = wb.create_sheet(title=hoja)
                excel_headers = [
                    "id", "entities_id", "is_recursive", "name", "ram", "serial", "otherserial", "contact", "contact_num", 
                    "users_id_tech", "groups_id_tech", "date_mod", "comment", "locations_id", "networks", "networks_id", 
                    "networkequipmenttypes_id", "networkequipmentmodels", "networkequipmentmodels_id", "manufacturers_id", "is_deleted", 
                    "is_template", "template_name", "users_id", "groups_id", "states_id", "ticket_tco", "is_dynamic", 
                    "uuid", "date_creation", "autoupdatesystems_id", "sysdescr", "cpu", "uptime", "last_inventory_update", 
                    "snmpcredentials_id", "links"
                ]
                ws.append(excel_headers)
            # Eliminar la hoja por defecto creada por Workbook
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(ruta)

    def crear_hoja_excel(self, wb, asset_type):
        if asset_type not in wb.sheetnames:
            ws = wb.create_sheet(title=asset_type)
            excel_headers = [
                    "id", "entities_id", "is_recursive", "name", "ram", "serial", "otherserial", "contact", "contact_num", 
                    "users_id_tech", "groups_id_tech", "date_mod", "comment", "locations_id", "networks", "networks_id", 
                    "networkequipmenttypes_id", "networkequipmentmodels", "networkequipmentmodels_id", "manufacturers_id", "is_deleted", 
                    "is_template", "template_name", "users_id", "groups_id", "states_id", "ticket_tco", "is_dynamic", 
                    "uuid", "date_creation", "autoupdatesystems_id", "sysdescr", "cpu", "uptime", "last_inventory_update", 
                    "snmpcredentials_id", "links"
            ]
            ws.append(excel_headers)
        else:
            ws = wb[asset_type]
            excel_headers = [cell.value for cell in ws[1]]
        return ws, excel_headers

    # --- Configuraciones ---
    # --- GLPI API ---

    def obtener_token_sesion(self):
        headers = {
            "Authorization": f"user_token {USER_TOKEN}",
            "App-Token": APP_TOKEN,
        }
        try:
            response = requests.get(f"{GLPI_URL}/initSession", headers=headers, verify=False)
            if response.status_code == 200:
                print("Sesión iniciada correctamente.")
                return response.json().get("session_token")
            else:
                print(f"Error al iniciar sesión: {response.status_code}")
                messagebox.showerror("Error", f"Error al iniciar sesión: {response.status_code}")
                return None
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            messagebox.showerror("Error", f"Error al conectar con la API de GLPI: {e}")
            return None
    
    def obtener_todos_los_network_equipment_glpi(self, session_token):
        """
        Obtiene todos los equipos de red en GLPI y los muestra de forma legible en la terminal.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        endpoint = "/NetworkEquipment"

        params = {"range": "0-999"}  # Asumiendo que tienes un rango para obtener varios registros

        response = requests.get(f"{GLPI_URL}{endpoint}", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            network_equipment_data = response.json()
            # Verificar si la respuesta contiene datos en formato esperado
            if isinstance(network_equipment_data, list):
                # Si ya es un arreglo, lo imprimimos directamente
                for equipment in network_equipment_data:
                    print(json.dumps(equipment, indent=4))
            else:
                # Si la respuesta está contenida en un diccionario, extraer los equipos
                equipment_list = network_equipment_data.get('data', [])
                for equipment in equipment_list:
                    print(json.dumps(equipment, indent=4))
        else:
            print("Error: No se pudieron obtener los equipos de red.")

    def obtener_network_equipment_glpi(self, session_token, asset_id):
        """
        Obtiene la información actual de un equipo de red en GLPI y la muestra de forma legible en la terminal.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        endpoint = "/NetworkEquipment"
        
        params = {"range": "0-999"}  # Asumiendo que tienes un rango para obtener varios registros

        # Obtener información de un equipo de red específico
        response = requests.get(f"{GLPI_URL}{endpoint}/{asset_id}", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            network_equipment_data = response.json()
            # Mostrar los datos de forma legible en la terminal
            print(json.dumps(network_equipment_data, indent=4))
        else:
            print(f"Error: No se pudo obtener el equipo de red con ID {asset_id}.")

    def obtener_todos_los_networks_id_glpi(self, session_token):
        """
        Obtiene todos los 'network_id' o "IP" de GLPI y muestra la respuesta de manera legible en la terminal.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        try:
            params = {"range": "0-999"}  # Obtener un rango de 0 a 999 para los networks
            
            endpoint = "/Network"

            # Hacer la solicitud GET a la API de GLPI para obtener los networks
            response = requests.get(f'{GLPI_URL}{endpoint}', headers=headers, params=params, verify=False)
            
            # Verificar si la respuesta fue exitosa (código 200)
            if response.status_code == 200:
                networks = response.json()
                # Imprimir la respuesta de manera legible
                print(json.dumps(networks, indent=4))
            else:
                print(f"Error al obtener los networks: {response.status_code} - {response.text}")
                print([])
        
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return []

    def obtener_id_de_networks_id(self, session_token, network):
        """
        Obtiene el ID de un 'network' en GLPI a partir de su nombre o dirección IP.
    
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        try:
            params = {"range": "0-999"}  # Obtener hasta 999 redes
            endpoint = "/Network"

            # Solicitud GET a la API de GLPI para obtener todos los networks
            response = requests.get(f'{GLPI_URL}{endpoint}', headers=headers, params=params, verify=False)

            if response.status_code == 200:
                networks = response.json()
                #print("Networks id", networks)

                # Buscar el network por nombre o IP
                for net in networks:
                    if net.get("name") == network:
                        print(f"Network encontrado: {json.dumps(net, indent=4)}")
                        #print(net.get("id"))
                        return net.get("id")  # Retornar el ID del network encontrado
 
                print(f"Network IP {network} no encontrado en GLPI.")
                return None

            else:
                print(f"Error al obtener los networks: {response.status_code} - {response.text}")
                return None

        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return None

    def obtener_todos_los_networkequipmentmodels_id_glpi(self, session_token):
        """
        Obtiene todos los 'networkequipmentmodels_id' de GLPI.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        try:
            params = {"range": "0-999"}  # Ajusta el rango según sea necesario

            # Hacer la solicitud GET a la API de GLPI para obtener los modelos de equipos de red
            response = requests.get(f'{GLPI_URL}/NetworkEquipmentModel', headers=headers, params=params, verify=False)
            
            # Verificar si la respuesta fue exitosa (código 200)
            if response.status_code == 200:
                networkequipmentmodels = response.json()
                # Extraer los networkequipmentmodels_id
                networkequipmentmodels_ids = [model['id'] for model in networkequipmentmodels]
                #print(f"Modelos de equipos de red encontrados: {networkequipmentmodels_ids}")
                # Imprimir la respuesta de manera legible
                print(json.dumps(networkequipmentmodels, indent=4))
                return networkequipmentmodels_ids
            else:
                print(f"Error al obtener los modelos de equipos de red: {response.status_code} - {response.text}")
                return []
        
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return []

    def crear_network_id_glpi(self, session_token, name, comment=""):
        """
        Crea un nuevo 'network_id' en GLPI.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        # Datos del nuevo 'network', dentro de un objeto 'input'
        new_network = {
            "input": [
                {
                    "name": name,
                    #"comment": comment
                }
            ]
        }
        
        try:
            # Hacer la solicitud POST a la API de GLPI para crear el nuevo 'network'
            response = requests.post(f'{GLPI_URL}/Network', headers=headers, json=new_network, verify=False)
            
            # Verificar si la respuesta fue exitosa (código 200 o 201)
            if response.status_code in [200, 201]:
                created_network = response.json()
                print(f"Nuevo Network creado: {json.dumps(created_network, indent=4)}")
                return 
            else:
                print(f"Error al crear el network: {response.status_code} - {response.text}")
                return None
            
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return None

    def crear_networkequipmentmodel_glpi(self, session_token, name, comment="", product_number="", weight=0, required_units=1, depth=1, power_connections=0, power_consumption=0, is_half_rack=0):
        """
        Crea un nuevo 'networkequipmentmodel' en GLPI.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        # Datos del nuevo 'networkequipmentmodel', dentro de un objeto 'input'
        new_model = {
            "input": [
                {
                    "name": name,
                    #"comment": comment,
                    #"product_number": product_number,
                    #"weight": weight,
                    #"required_units": required_units,
                    #"depth": depth,
                    #"power_connections": power_connections,
                    #"power_consumption": power_consumption,
                    #"is_half_rack": is_half_rack
                }
            ]
        }
        
        try:
            # Hacer la solicitud POST a la API de GLPI para crear el nuevo 'networkequipmentmodel'
            response = requests.post(f'{GLPI_URL}/NetworkEquipmentModel', headers=headers, json=new_model, verify=False)
            
            # Verificar si la respuesta fue exitosa (código 200 o 201)
            if response.status_code in [200, 201]:
                created_model = response.json()
                print(f"Nuevo NetworkEquipmentModel creado: {json.dumps(created_model, indent=4)}")
                return created_model
            else:
                print(f"Error al crear el NetworkEquipmentModel: {response.status_code} - {response.text}")
                return None
        
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return None

    def obtener_id_de_networkequipmentmodels(self, session_token, networkequipmentmodel):
        """
        Obtiene el ID de un 'networkequipmentmodel' en GLPI a partir de su nombre o modelo.
        
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        try:
            params = {"range": "0-999"}  # Obtener hasta 999 modelos de equipos
            endpoint = "/NetworkEquipmentModel"

            # Solicitud GET a la API de GLPI para obtener todos los modelos de equipos
            response = requests.get(f'{GLPI_URL}{endpoint}', headers=headers, params=params, verify=False)

            if response.status_code == 200:
                networkequipmentmodels_data = response.json()

                # Buscar el modelo de equipo por nombre o modelo
                for model in networkequipmentmodels_data:
                    if model.get("name") == networkequipmentmodel: #or model.get("product_number") == networkequipmentmodel:
                        print(f"Modelo de equipo encontrado: {json.dumps(model, indent=4)}")
                        return model.get("id")  # Retornar el ID del modelo de equipo encontrado

                print("Modelo de equipo no encontrado en GLPI.")
                return None

            else:
                print(f"Error al obtener los modelos de equipo: {response.status_code} - {response.text}")
                return None

        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return None

    # --- Metodo ---

    def registro_offline(self):
        qr_data = scanner.escanear_qr_con_celular_firstDetection()

        # Si se escanea correctamente, agregamos el equipo al Excel
        if qr_data:
            self.agregar_equipo_a_excel(ruta_excel, qr_data)

    def multiples_registros_offline(self):
        qr_data = scanner.escanear_qr_con_celular()
        #print(qr_data)
        # Si se escanea correctamente, agregamos el equipo al Excel
        if qr_data:
            self.agregar_equipos_a_excel(qr_data)

    def registro_online(self): 
        qr_data = scanner.escanear_qr_con_celular_firstDetection()
        #print(qr_data)

        session_token = self.obtener_token_sesion()
        # Si se escanea correctamente, agregamos el equipo al Excel
        if qr_data:
            self.agregar_equipo_a_GLPI(session_token, qr_data)

    def multiples_registros_online(self):
        qr_data = scanner.escanear_qr_con_celular()
        #print(qr_data)

        session_token = self.obtener_token_sesion()
        # Si se escanea correctamente, agregamos el equipo al Excel
        if qr_data:
            self.agregar_multiples_equipos_a_GLPI(session_token, qr_data)

    def sincronizacion_asincrona(self):
        print("Iniciando sincronización asincrónica...")
        session_token = self.obtener_token_sesion()
        if session_token:
            print("Token de sesión obtenido correctamente.")
            self.agregar_equipo_desde_excel_a_glpi(session_token, ruta_excel)
        else:
            print("No se pudo obtener el token de sesión de GLPI.")
            messagebox.showerror("Error", "No se pudo obtener el token de sesión de GLPI.")

    def agregar_equipo_desde_excel_a_glpi(self, session_token, ruta_excel): 
        """
        Agrega equipos de red desde un archivo Excel a GLPI.
        """

        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Abrir el archivo Excel
        wb = load_workbook(ruta_excel)
        ws = wb["NetworkEquipment new"]

        filas_a_eliminar = []  # Lista para almacenar los índices de las filas a eliminar
        tiene_datos = False  # Bandera para verificar si hay datos en el Excel

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  
            # Extraer los datos de cada fila
            name = row[3]  # 'name' está en la columna 4 (index 3)
            serial = row[5]  # 'serial' está en la columna 6 (index 5)
            ip = row[14]  # 'networks' está en la columna 15 (index 14)
            networkequipmentmodel = row[17]  # 'networkequipmentmodels' está en la columna 18 (index 17)

            if any([name, serial, ip, networkequipmentmodel]):  # Verifica si la fila tiene datos
                tiene_datos = True

            # Verificar si el equipo ya está registrado en GLPI
            qr_info = {"name": name, "serial": serial, "ip": ip, "model": networkequipmentmodel}
            existencia = self.verificar_equipo_existente_glpi(session_token, qr_info)

            if existencia:
                # Si el equipo ya existe en GLPI, marcar la fila para eliminar
                filas_a_eliminar.append(idx)
                print(f"El equipo con nombre {name} y serial {serial} ya está registrado en GLPI. Se eliminará del Excel de equipos nuevos...")
                messagebox.showinfo("Información", f"El equipo {name} ya está registrado en GLPI. Se eliminará del Excel de equipos nuevos...")
            else:
                # Obtener ID de la red (si no existe, crearlo)
                networks_id = self.obtener_id_de_networks_id(session_token, ip)
                if not networks_id:
                    messagebox.showinfo("Información", f"El IP {ip} no existe en GLPI. Se procederá a crear en GLPI.")
                    self.crear_network_id_glpi(session_token, ip)
                    networks_id = self.obtener_id_de_networks_id(session_token, ip)

                # Obtener ID del modelo de equipo de red (si no existe, crearlo)
                networkequipmentmodels_id = self.obtener_id_de_networkequipmentmodels(session_token, networkequipmentmodel)
                if not networkequipmentmodels_id:
                    messagebox.showinfo("Información", f"El modelo {networkequipmentmodel} no existe en GLPI. Se procederá a agregar.")
                    self.crear_networkequipmentmodel_glpi(session_token, networkequipmentmodel)
                    networkequipmentmodels_id = self.obtener_id_de_networkequipmentmodels(session_token, networkequipmentmodel)

                # Estructura del equipo para registrar en GLPI
                nuevo_equipo = {
                    "input": [
                        {
                            "name": name,
                            "serial": serial,
                            "networks_id": networks_id,
                            "networkequipmentmodels_id": networkequipmentmodels_id
                        }
                    ]
                }

                try:
                    # Hacer la solicitud POST a la API de GLPI para registrar el nuevo equipo
                    response = requests.post(f'{GLPI_URL}/NetworkEquipment', headers=headers, json=nuevo_equipo, verify=False)

                    # Verificar si la respuesta fue exitosa (código 200 o 201)
                    if response.status_code in [200, 201]:
                        created_equipment = response.json()
                        print(f"Equipo de red registrado en GLPI: {json.dumps(created_equipment, indent=4)}")
                        messagebox.showinfo("Información", f"Equipo {name} agregado correctamente a GLPI desde el Excel de equipos nuevos...")
                    else:
                        print(f"Error al registrar el equipo en GLPI: {response.status_code} - {response.text}")
                        messagebox.showerror("Error", f"No se pudo agregar el equipo {name} a GLPI.")
                except requests.exceptions.RequestException as e:
                    print(f"Error al conectar con la API de GLPI: {e}")
                    messagebox.showerror("Error", f"Error de conexión con la API de GLPI para el equipo {name}.")

        # Eliminar filas en orden inverso para evitar problemas de desplazamiento de índices
        for i in sorted(filas_a_eliminar, reverse=True):
            ws.delete_rows(i)

        # Guardar el archivo Excel con los cambios
        wb.save(ruta_excel)

        # Verificar si el Excel quedó vacío
        if not tiene_datos or ws.max_row == 1:
            print("El Excel de equipos nuevos está vacío desde antes o después de la eliminación.")
            messagebox.showwarning("Advertencia", "El Excel de equipos nuevos está vacío después de la eliminación.")

    def agregar_equipo_a_excel(self, ruta_excel, qr_data):
        with self.lock_excel:
            # Procesar los datos del QR
            qr_info = self.procesar_qr_data(qr_data)
            
            # Abrir el archivo Excel
            wb = load_workbook(ruta_excel)
            ws = wb["NetworkEquipment new"]

            # Verificar si el equipo ya está en el Excel por 'name' o 'serial'
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3] == qr_info.get("name") or row[5] == qr_info.get("serial"):
                    messagebox.showinfo("Información", "El equipo ya está registrado en el Excel. No se agregará.")
                    return  # Evita registros duplicados

            # Crear una nueva fila con los datos del QR
            nueva_fila = [
                None,  # id
                None,  # entities_id
                None,  # is_recursive
                qr_info.get("name"),  # name (NS del QR)
                None,  # ram
                qr_info.get("serial"),  # serial (SN del QR)
                None,  # otherserial
                None,  # contact
                None,  # contact_num
                None,  # users_id_tech
                None,  # groups_id_tech
                None,  # date_mod
                None,  # comment
                None,  # locations_id
                qr_info.get("ip"),  # networks
                qr_info.get("networks_id", None),  # networks_id (si existe)
                None,  # networkequipmenttypes_id
                qr_info.get("model"),  # networkequipmentmodels
                qr_info.get("networkequipmentmodels_id", None),  # networkequipmentmodels_id (si existe)
                None,  # manufacturers_id
                None,  # is_deleted
                None,  # is_template
                None,  # template_name
                None,  # users_id
                None,  # groups_id
                None,  # states_id
                None,  # ticket_tco
                None,  # is_dynamic
                None,  # uuid
                None,  # date_creation
                None,  # autoupdatesystems_id
                None,  # sysdescr
                None,  # cpu
                None,  # uptime
                None,  # last_inventory_update
                None,  # snmpcredentials_id
                None   # links
            ]

            # Añadir la nueva fila y guardar el archivo
            ws.append(nueva_fila)
            wb.save(ruta_excel)
            messagebox.showinfo("Información", "Equipo agregado correctamente.")

    def agregar_equipos_a_excel(self, lista_qr_data):
        """
        Agrega múltiples equipos escaneados a un archivo Excel, evitando duplicados.
        """
        try:
            # Obtener el bloqueo antes de acceder al archivo
            with self.lock_excel:
                # Cargar el archivo Excel
                wb = load_workbook(ruta_excel)
                ws = wb["NetworkEquipment new"]

                # Obtener valores actuales para evitar duplicados
                equipos_registrados = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[3]:  # 'name' en la columna correspondiente
                        equipos_registrados.add(row[3])
                    if row[5]:  # 'serial' en la columna correspondiente
                        equipos_registrados.add(row[5])

                # Procesar y agregar cada QR
                nuevos_equipos = 0
                for qr_data in lista_qr_data:
                    qr_info = self.procesar_qr_data(qr_data)
                    name = qr_info.get("name")
                    serial = qr_info.get("serial")

                    # Verificar duplicados
                    if name in equipos_registrados or serial in equipos_registrados:
                        messagebox.showinfo("Información", f"El equipo {name} ({serial}) ya está registrado. No se agregará.")
                        continue

                    # Agregar equipo
                    nueva_fila = [
                        None, None, None, name, None, serial, None, None, None, None, None, None, None, None,
                        qr_info.get("ip"), qr_info.get("networks_id", None), None, qr_info.get("model"),
                        qr_info.get("networkequipmentmodels_id", None), None, None, None, None, None, None,
                        None, None, None, None, None, None, None, None, None, None
                    ]
                    ws.append(nueva_fila)
                    equipos_registrados.add(name)
                    equipos_registrados.add(serial)
                    nuevos_equipos += 1

                # Guardar cambios
                wb.save(ruta_excel)
                messagebox.showinfo("Información", f"Se agregaron {nuevos_equipos} equipos nuevos correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar Excel: {str(e)}")

    def agregar_equipo_a_GLPI(self, session_token, qr_data):
        """
        Registra un nuevo equipo de red en GLPI a partir de los datos extraídos del QR.
        """
        with self.lock_glpi:
            headers = {
                "Content-Type": "application/json",
                "Session-Token": session_token,
                "App-Token": APP_TOKEN
            }

            # Procesar los datos del QR
            qr_info = self.procesar_qr_data(qr_data)

            name= qr_info.get("name")
            serial = qr_info.get("serial")
            ip = str(qr_info.get("ip"))
            networks_id = self.obtener_id_de_networks_id(session_token, ip)

            if not networks_id:
                messagebox.showinfo("Información", f"El IP {ip} no existe en GLPI. Se procedera a crear en GLPI")
                self.crear_network_id_glpi(session_token, ip)
                networks_id = self.obtener_id_de_networks_id(session_token, ip)

            networkequipmentmodel =  qr_info.get("model")
            networkequipmentmodels_id = self.obtener_id_de_networkequipmentmodels(session_token, networkequipmentmodel)

            if not networkequipmentmodels_id:    
                messagebox.showinfo("Información", f"El modelo{networkequipmentmodel} no existe en GLPI. Se procedera a agregar")
                self.crear_networkequipmentmodel_glpi(session_token, networkequipmentmodel)
                networkequipmentmodels_id = self.obtener_id_de_networkequipmentmodels(session_token, networkequipmentmodel)
            
            # Estructura del equipo de red a registrar en GLPI
            nuevo_equipo = {
                "input": [
                    {
                        "name": name,  # Nombre del equipo
                        "serial": serial,  # Número de serie
                        "networks_id": networks_id,  # Dirección IP
                        "networkequipmentmodels_id": networkequipmentmodels_id,  # ID del modelo de equipo de red
                        #"comment": qr_info.get("comment", ""),  # Comentario opcional
                        #"locations_id": qr_info.get("locations_id", ""),  # Ubicación (si aplica)
                        #"manufacturers_id": qr_info.get("manufacturers_id", ""),  # Fabricante (si aplica)
                        #"users_id": qr_info.get("users_id", "")  # Usuario asignado (si aplica)
                    }
                ]
            }

            try:
                existencia = self.verificar_equipo_existente_glpi(session_token, qr_info)

                if not existencia:
                    # Hacer la solicitud POST a la API de GLPI para registrar el nuevo equipo
                    response = requests.post(f'{GLPI_URL}/NetworkEquipment', headers=headers, json=nuevo_equipo, verify=False)

                    # Verificar si la respuesta fue exitosa (código 200 o 201)
                    if response.status_code in [200, 201]:
                        created_equipment = response.json()
                        print(f"Equipo de red registrado en GLPI: {json.dumps(created_equipment, indent=4)}")
                        messagebox.showinfo("Información", f"Equipo de red registrado {created_equipment["name"]} en GLPI")
                        #self.agregar_equipo_a_excel(ruta_excel, qr_data)
                        return created_equipment
                    else:
                        print(f"Error al registrar el equipo en GLPI: {response.status_code} - {response.text}")
                        return None
                else: 
                    messagebox.showinfo("Información", f"El equipo con nombre {name} y serial {serial} ya está registrado en GLPI. No se agregara...")
                    return None

            except requests.exceptions.RequestException as e:
                print(f"Error al conectar con la API de GLPI: {e}")
                return None

    def agregar_multiples_equipos_a_GLPI(self, session_token, lista_qr_data):
        """
        Registra múltiples equipos de red en GLPI a partir de los datos extraídos de una lista de QR.
        """
        for qr_data in lista_qr_data:
            try:
                # Llamar a la función agregar_equipo_a_GLPI para cada equipo
                created_equipment = self.agregar_equipo_a_GLPI(session_token, qr_data)

                if created_equipment:
                    print(f"Equipo registrado correctamente: {created_equipment['input'][0]['name']}")
                else:
                    print(f"Error al registrar el equipo con QR: {qr_data}")
            except Exception as e:
                print(f"Error procesando el QR {qr_data}: {str(e)}")

    def procesar_qr_data(self, qr_data):
        # Separar los datos por líneas
        lineas = qr_data.split('\n')
        
        # Crear un diccionario para almacenar los valores
        qr_info = {}
        
        # Buscar las claves y extraer los valores
        for linea in lineas:
            if linea.startswith("NS:"):
                qr_info["name"] = linea.split("NS:")[1].strip()
            elif linea.startswith("IP:"):
                qr_info["ip"] = linea.split("IP:")[1].strip()
            elif linea.startswith("SN:"):
                qr_info["serial"] = linea.split("SN:")[1].strip()
            elif linea.startswith("MODEL:"):
                qr_info["model"] = linea.split("MODEL:")[1].strip()
        return qr_info

    def verificar_equipo_existente_glpi(self, session_token, qr_info):
        """
        Verifica si un equipo ya existe en GLPI por 'name' o 'serial'.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        # Parámetros para búsqueda por nombre o serial
        params = {
            "searchText": qr_info.get("name"),  # Búsqueda por nombre
            "range": "0-999",  # Limitar a 999 resultados
        }

        endpoint = "/NetworkEquipment"

        try:
            # Realizar solicitud para buscar el equipo por 'name'
            response_check = requests.get(f'{GLPI_URL}{endpoint}', headers=headers, params=params, verify=False)

            if response_check.status_code == 200:
                equipment_data = response_check.json()
                
                # Verificar si la respuesta contiene equipos
                if isinstance(equipment_data, list):
                    for equipment in equipment_data:
                        # Verificar si el nombre o el número de serie ya existe en GLPI
                        if equipment.get("name") == qr_info.get("name") and equipment.get("serial") == qr_info.get("serial"):
                            print(f"El equipo con nombre '{qr_info.get('name')}' y serial '{qr_info.get('serial')}' ya está registrado en GLPI.")
                            return True  # El equipo ya está registrado
                else:
                    print("La respuesta de GLPI no contiene equipos válidos.")
                    return False
            else:
                print(f"Error al verificar el equipo en GLPI: {response_check.status_code} - {response_check.text}")
                return False
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            return False

    def obtener_todos_los_network_equipment_glpi_a_excel(self, session_token):
        """
        Obtiene todos los equipos de red en GLPI y los guarda en la hoja "NetworkEquipment" del archivo Excel.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        endpoint = f"{GLPI_URL}/NetworkEquipment"
        params = {"range": "0-999"}  # Ajustar el rango según necesidad

        response = requests.get(endpoint, headers=headers, params=params, verify=False)
        
        if response.status_code == 200:
            network_equipment_data = response.json()
            
            # Verificar si la respuesta es una lista o un diccionario
            equipment_list = network_equipment_data if isinstance(network_equipment_data, list) else network_equipment_data.get('data', [])
            
            if not equipment_list:
                print("No se encontraron equipos de red.")
                return
            
            # Definir las columnas requeridas
            columnas = [
                "id", "entities_id", "is_recursive", "name", "ram", "serial", "otherserial", "contact", "contact_num", 
                "users_id_tech", "groups_id_tech", "date_mod", "comment", "locations_id", "networks_id", 
                "networkequipmenttypes_id", "networkequipmentmodels_id", "manufacturers_id", "is_deleted", 
                "is_template", "template_name", "users_id", "groups_id", "states_id", "ticket_tco", "is_dynamic", 
                "uuid", "date_creation", "autoupdatesystems_id", "sysdescr", "cpu", "uptime", "last_inventory_update", 
                "snmpcredentials_id", "links"
            ]
            
            # Convertir a DataFrame asegurando que solo se extraigan las columnas requeridas
            df = pd.DataFrame(equipment_list)
            df = df[columnas]  # Seleccionar solo las columnas especificadas
            
            # Verificar si el archivo ya existe
            if os.path.exists(ruta_excel):
                with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name="NetworkEquipment", index=False)
            else:
                with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="NetworkEquipment", index=False)
            
            print(f"Equipos de red guardados en la hoja 'NetworkEquipment' del archivo {ruta_excel}")
        else:
            print(f"Error: No se pudieron obtener los equipos de red. Código {response.status_code}")

class QRScanner:
    def __init__(self):
        self.qr_escaneados = set()  # Almacena los QR únicos escaneados
        self.lock = threading.Lock()  # Lock para controlar el acceso a la cámara

    def es_codigo_valido(self, qr_data):
        # Normalizar el texto eliminando espacios extra y líneas vacías
        qr_data_limpio = "\n".join([line.strip() for line in qr_data.splitlines() if line.strip()])
        
        print("Texto procesado antes de la validación:")
        print(repr(qr_data_limpio))  # Muestra caracteres invisibles como '\n'
        
        # Expresión regular corregida
        patron_valido = r'^NS:[A-Za-z0-9\-.]+\nIP:\d{1,3}(\.\d{1,3}){3}\nSN:[A-Za-z0-9]+\nMODEL:[A-Za-z0-9\s]+$'
        
        # Validar el código QR contra el patrón
        if re.match(patron_valido, qr_data_limpio, re.MULTILINE):
            return "valido"
        else:
            return "invalido"

    def escanear_qr_con_celular(self):
        """Escanea múltiples códigos QR y evita duplicados. Presiona 'q' para salir."""
        try:
            respuesta = messagebox.askokcancel("Confirmación", "¿Desea activar la cámara para escanear el QR?")
            if not respuesta:
                return None

            messagebox.showinfo("Información", "Escaneando con la cámara del celular. Presiona 'q' para salir.")

            result_queue = queue.Queue()

            def run_capture(result_queue):
                with self.lock:
                    cap = cv2.VideoCapture(IP_CAM_URL)
                    if not cap.isOpened():
                        result_queue.put(("error", "No se pudo acceder a la cámara del celular."))
                        return
                    
                    while True:
                        ret, frame = cap.read()
                        if not ret:
                            result_queue.put(("error", "Error al obtener el cuadro de la cámara."))
                            break

                        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        qr_codes = decode(gray_frame)

                        for qr in qr_codes:
                            qr_data = qr.data.decode('utf-8').strip()

                            if qr_data in self.qr_escaneados:
                                continue  # Si ya se escaneó, lo ignora

                            if self.es_codigo_valido(qr_data) == "valido":
                                self.qr_escaneados.add(qr_data)
                                result_queue.put(("info", f"QR válido detectado: \n{qr_data}"))
                                result_queue.put(("data", qr_data))

                        cv2.imshow("Escaneando QR con celular", frame)

                        if cv2.waitKey(1) & 0xFF == ord('q'):
                            result_queue.put(("close", None))
                            break

                        if cv2.getWindowProperty("Escaneando QR con celular", cv2.WND_PROP_VISIBLE) < 1:
                            result_queue.put(("close", None))
                            break

                    cap.release()
                    cv2.destroyAllWindows()

            capture_thread = threading.Thread(target=run_capture, args=(result_queue,))
            capture_thread.start()

            while True:
                try:
                    msg_type, msg_content = result_queue.get(timeout=1)
                    if msg_type == "info":
                        messagebox.showinfo("Información", msg_content)
                    elif msg_type == "error":
                        messagebox.showerror("Error", msg_content)
                        return None
                    elif msg_type == "data":
                        print(f"QR almacenado: {msg_content}")
                    elif msg_type == "close":
                        print("Escaneo finalizado.")
                        return list(self.qr_escaneados)  # Devuelve todos los QR escaneados
                except queue.Empty:
                    continue

        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")
            cv2.destroyAllWindows()

    def escanear_qr_con_celular_firstDetection(self):
        try:
            # Mostrar cuadro de diálogo de confirmación
            respuesta = messagebox.askokcancel("Confirmación", "¿Desea activar la cámara para escanear el QR?")
            if not respuesta:
                return None

            # Mostrar información adicional después de la confirmación
            messagebox.showinfo("Información", "Usando la cámara del celular. Presiona 'q' para salir.")

            def run_capture(result_queue):
                ip_cam_url = IP_CAM_URL  # Cambiar por la URL de la cámara IP
                camera_open = True

                try:
                    with self.lock:  # Asegura que solo una función pueda acceder a la cámara a la vez
                        cap = cv2.VideoCapture(ip_cam_url)
                        if not cap.isOpened():
                            result_queue.put(("error", "No se pudo acceder a la cámara del celular. Reintentalo..."))
                            cap.release()
                            cv2.destroyAllWindows()
                            return

                        while camera_open:
                            ret, frame = cap.read()
                            if not ret:
                                result_queue.put(("error", "Error al obtener el cuadro de la cámara. Reintentando conexión..."))
                                break  # Sale del bucle interno para reintentar la conexión

                            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                            qr_codes = decode(gray_frame)

                            for qr in qr_codes:
                                qr_data = qr.data.decode('utf-8')
                                flag = self.es_codigo_valido(qr_data)
                                if flag == "valido":
                                    result_queue.put(("info", f"Código QR {flag}: \n{qr_data}"))
                                    result_queue.put(("data", qr_data))
                                    camera_open = False
                                    cap.release()
                                    cv2.destroyAllWindows()
                                    return
                                elif flag == "invalido":
                                    continue

                            cv2.imshow("Escaneando QR con celular", frame)

                            if cv2.waitKey(1) & 0xFF == ord('q'):
                                camera_open = False
                                cap.release()
                                cv2.destroyAllWindows()
                                result_queue.put(("close", None))
                                return

                            # Verificar si la ventana fue cerrada
                            if cv2.getWindowProperty("Escaneando QR con celular", cv2.WND_PROP_VISIBLE) < 1:
                                camera_open = False
                                cap.release()
                                cv2.destroyAllWindows()
                                result_queue.put(("close", None))
                                return

                        cap.release()
                        cv2.destroyAllWindows()
                except Exception as e:
                    result_queue.put(("error", f"Se produjo un error inesperado: {str(e)}"))
                    cap.release()
                    cv2.destroyAllWindows()

            result_queue = queue.Queue()
            capture_thread = threading.Thread(target=run_capture, args=(result_queue,))
            capture_thread.start()

            while True:
                try:
                    msg_type, msg_content = result_queue.get(timeout=1)
                    if msg_type == "info":
                        messagebox.showinfo("Información", msg_content)
                    elif msg_type == "error":
                        messagebox.showerror("Error", msg_content)
                        return None
                    elif msg_type == "data":
                        return msg_content
                    elif msg_type == "close":
                        return None
                except queue.Empty:
                    continue
        except Exception as e:
            result_queue.put(("error", f"Se produjo un error inesperado: {str(e)}"))
            cv2.destroyAllWindows()


if __name__ == "__main__":
    root = tk.Tk()
    scanner = QRScanner()
    app = NetworkEquipment(root)
    root.mainloop()
